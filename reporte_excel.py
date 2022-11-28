from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart, PieChart
from openpyxl.styles import Font
import matplotlib.pyplot as plt
import xlwings as xw
from process_data import usar_reportes
import seaborn as sns
import pandas as pd
import signal
import sys


def handler(signal, frame):
    entrada = input('Has pulsado ctrl+c. Quieres salir del programa(Y/n): ')
    if entrada == 'Y':
        print('Saliendo de manera controlada')
        sys.exit(1)
    else:
        print('Continuando con la ejecucion')


def extract():
    mes_año, revenues_per_week, revenues, pizzas_vendidas, revenues_por_pedido, comida_cena, dia_semana, cantidad_pizzas_pedido, pizzas_pedidas = usar_reportes()
    comida_cena = pd.DataFrame({'Comida': [comida_cena[False]], 'Cena': [
        comida_cena[True]]}, index=['Probability'])
    ingredientes = pd.read_csv('compra_semanal_2017.csv')
    a = ingredientes.sum().sort_values(ascending=False)[
        :20]/ingredientes.sum().sum()
    # Los 20 ingredientes mas usados.
    ingredientes_mas_usados = pd.DataFrame(a.to_dict(), index=['Probabilidad'])
    dia_semana = pd.DataFrame(dia_semana.to_dict(), index=['Probabilidad'])
    mes_año = pd.DataFrame(mes_año.to_dict(), index=['Probabilidad'])
    return mes_año, revenues_per_week, revenues, \
        pizzas_vendidas, revenues_por_pedido, \
        comida_cena, dia_semana, \
        cantidad_pizzas_pedido, pizzas_pedidas, ingredientes_mas_usados


def create_book(comida, ingredientes_mas_usados, dia_semana, mes_año):
    ingredientes = pd.read_csv('compra_semanal_2017.csv')
    a = ingredientes.sum().sort_values(ascending=False)[
        :20]/ingredientes.sum().sum()
    # Los 20 ingredientes mas usados.
    b = pd.DataFrame(a.to_dict(), index=['Probabilidad'])
    with pd.ExcelWriter('Reporte_excel.xlsx') as writer:
        dia_semana.to_excel(writer, sheet_name='Reporte Pedidos', startrow=3)
        mes_año.to_excel(writer, sheet_name='Reporte Pedidos', startrow=5)
        comida.to_excel(writer,
                        sheet_name='Reporte Pedidos', startrow=0)
        ingredientes.to_excel(writer,
                              sheet_name='Reporte Ingredientes', index=False)
        ingredientes_mas_usados.to_excel(writer,
                                         sheet_name='Reporte Ingredientes', startrow=54)

    libro = load_workbook("Reporte_excel.xlsx")
    return libro, a


def pedidos(libro):
    reporte_pedidos = libro['Reporte Pedidos']
    pie = PieChart()
    data = Reference(reporte_pedidos, min_col=2,
                     max_col=3, min_row=2, max_row=2)
    labels = Reference(reporte_pedidos, min_col=2,
                       max_col=3, min_row=1, max_row=1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Probability for dinner and lunch"
    reporte_pedidos.add_chart(pie, "D1")
    values = Reference(reporte_pedidos, min_col=2,
                       max_col=8, min_row=4, max_row=5)
    print(values)
    grafico2 = BarChart()
    grafico2.add_data(values, titles_from_data=True)
    grafico2.title = 'Pizzas vendidad por dia de la semana'
    grafico2.style = 10
    grafico2.x_axis.title = 'Dia de la semana'
    grafico2.y_axis.title = 'Pizzas'
    reporte_pedidos.add_chart(grafico2, "K1")
    ####################################################
    values = Reference(reporte_pedidos, min_col=2,
                       max_col=13, min_row=6, max_row=7)
    grafico1 = BarChart()
    grafico1.add_data(values, titles_from_data=True)
    grafico1.title = 'Pizzas vendidas por mes'
    grafico1.style = 10
    grafico1.x_axis.title = 'Mes'
    grafico1.y_axis.title = 'Pizzas'
    reporte_pedidos.add_chart(grafico1, "K1")
    reporte_pedidos['D1'] = 'Pizza type'
    reporte_pedidos['A4'] = 'Week day'
    reporte_pedidos['A12'] = 'Month'
    reporte_pedidos['D1'].font = Font(bold=True, color='FF0000')
    reporte_pedidos['A4'].font = Font(bold=True, color='0000FF')
    reporte_pedidos['A12'].font = Font(bold=True, color='99004C')
    safe_book(libro)


def ingredientes(libro):
    reporte_ingredientes = libro['Reporte Ingredientes']
    values = Reference(reporte_ingredientes, min_col=2,
                       max_col=21, min_row=55, max_row=56)
    grafico1 = BarChart()
    grafico1.add_data(values, titles_from_data=True)
    grafico1.title = f'Ingredientes más usados: Representan el {int(a.sum()*100)}%'
    grafico1.style = 10
    grafico1.x_axis.title = 'Ingredientes'
    grafico1.y_axis.title = 'Probabilidad de usar el ingrediente'
    reporte_ingredientes.add_chart(grafico1, "K1")
    safe_book(libro)


def safe_book(libro):
    libro.save('Reporte_excel.xlsx')


def ejecutivo(libro, revenues, pizzas_vendidas, cantidad_pizzas_pedido, revenues_per_week, pizzas_pedidas, a):
    reporte_ejecutivo = libro.create_sheet('Reporte Ejecutivo')
    reporte_ejecutivo['A1'] = 'Revenues'
    reporte_ejecutivo['A2'] = int(revenues)
    reporte_ejecutivo['B1'] = 'Pizzas Vendidas'
    reporte_ejecutivo['B2'] = int(pizzas_vendidas)
    reporte_ejecutivo['C1'] = 'Pedidos hechos'
    reporte_ejecutivo['C2'] = 21350
    reporte_ejecutivo['D1'] = 'Media de pizzas por pedidos'
    reporte_ejecutivo['D2'] = int(cantidad_pizzas_pedido)
    safe_book(libro)
    ###################################
    fig = plt.figure(figsize=(8, 4))
    plt.plot(revenues_per_week)
    plt.title('Revenues per week')
    wb = xw.Book('Reporte_excel.xlsx')
    sheet = wb.sheets['Reporte Ejecutivo']
    sheet.pictures.add(fig, name='Revenues per week', update=True)
    ####################################
    fig2 = plt.figure(figsize=(4, 3))
    sns.barplot(x=pizzas_pedidas[:10].index,
                y=pizzas_pedidas[:10], orient='v', palette='rocket_r')
    plt.title('Most sold pizzas')
    plt.xticks(rotation=70)
    sheet.pictures.add(fig2, name='Most Sold Pizzas', update=True)


if __name__ == '__main__':
    signal.signal(signal.SIGINT, handler)
    mes_año, revenues_per_week, revenues, \
        pizzas_vendidas, revenues_por_pedido, \
        comida_cena, dia_semana, \
        cantidad_pizzas_pedido, pizzas_pedidas, ingredientes_mas_usados = extract()
    libro, a = create_book(
        comida_cena, ingredientes_mas_usados, dia_semana, mes_año)
    pedidos(libro)
    ingredientes(libro)
    ejecutivo(libro, revenues, pizzas_vendidas,
              cantidad_pizzas_pedido, revenues_per_week, pizzas_pedidas, a)  # Le pasamos mas cosas porque tenemos que hace run plot.
